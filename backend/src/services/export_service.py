"""
Export Service.

Generates JSON, CSV, Excel (.xlsx), and ZIP exports
from extracted document data.
"""

import csv
import io
import json
import logging
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class ExportService:
    """Handles all export format generation."""

    # ── JSON ──────────────────────────────────────────────────────────────

    def to_json(self, documents: list[dict]) -> bytes:
        """Export documents as a formatted JSON byte string."""
        payload = {
            "export_date": datetime.utcnow().isoformat(),
            "total": len(documents),
            "documents": documents,
        }
        return json.dumps(payload, indent=2, default=str).encode("utf-8")

    # ── CSV ───────────────────────────────────────────────────────────────

    def to_csv(self, documents: list[dict]) -> bytes:
        """
        Flatten extracted fields into a CSV.
        One row per document; nested objects are JSON-encoded.
        """
        if not documents:
            return b""

        # Collect all field keys across documents
        all_keys = set()
        for doc in documents:
            fields = doc.get("extraction_result", {}).get("fields", {})
            all_keys.update(fields.keys())

        base_cols = ["id", "original_filename", "document_type", "status", "created_at"]
        field_cols = sorted(all_keys)
        all_cols = base_cols + field_cols

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=all_cols, extrasaction="ignore")
        writer.writeheader()

        for doc in documents:
            fields = doc.get("extraction_result", {}).get("fields", {})
            row = {col: doc.get(col, "") for col in base_cols}
            for key in field_cols:
                val = fields.get(key, "")
                # Flatten nested structures
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, default=str)
                row[key] = val
            writer.writerow(row)

        return output.getvalue().encode("utf-8")

    # ── Excel ─────────────────────────────────────────────────────────────

    def to_xlsx(self, documents: list[dict]) -> bytes:
        """
        Export as Excel workbook.
        - Sheet 'Summary': one row per document
        - Per-document-type sheets with type-specific columns
        - Tables get their own sheet per document
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed")
            return self.to_csv(documents)

        wb = Workbook()

        # ── Summary sheet ─────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        header_style = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2C3E50")
        headers = ["ID", "Filename", "Type", "Status", "Confidence", "Created At"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=h)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, doc in enumerate(documents, 2):
            result = doc.get("extraction_result") or {}
            ws_summary.cell(row=row_idx, column=1, value=doc.get("id", ""))
            ws_summary.cell(row=row_idx, column=2, value=doc.get("original_filename", ""))
            ws_summary.cell(row=row_idx, column=3, value=doc.get("document_type", ""))
            ws_summary.cell(row=row_idx, column=4, value=doc.get("status", ""))
            ws_summary.cell(row=row_idx, column=5, value=result.get("overall_confidence", ""))
            ws_summary.cell(row=row_idx, column=6, value=str(doc.get("created_at", "")))

        self._auto_width(ws_summary)

        # ── Per-type sheets ────────────────────────────────────────────────
        by_type: dict[str, list] = {}
        for doc in documents:
            dt = doc.get("document_type") or "unknown"
            by_type.setdefault(dt, []).append(doc)

        for doc_type, type_docs in by_type.items():
            sheet_name = doc_type[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)

            # Collect all field keys for this type
            all_field_keys: set = set()
            for doc in type_docs:
                fields = (doc.get("extraction_result") or {}).get("fields", {})
                all_field_keys.update(fields.keys())

            col_headers = ["Filename"] + sorted(all_field_keys)
            for col_idx, h in enumerate(col_headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_style
                cell.fill = header_fill

            for row_idx, doc in enumerate(type_docs, 2):
                fields = (doc.get("extraction_result") or {}).get("fields", {})
                ws.cell(row=row_idx, column=1, value=doc.get("original_filename", ""))
                for col_idx, key in enumerate(sorted(all_field_keys), 2):
                    val = fields.get(key, "")
                    if isinstance(val, (list, dict)):
                        val = json.dumps(val, default=str)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            self._auto_width(ws)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ── ZIP ───────────────────────────────────────────────────────────────

    def to_zip(
        self,
        documents: list[dict],
        include_images: bool = False
    ) -> bytes:
        """
        ZIP archive containing:
        - summary.json (all documents)
        - summary.csv
        - per_document/<id>.json (individual)
        - images/<id>.jpg (if include_images)
        """
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            # Summary files
            zf.writestr("summary.json", self.to_json(documents).decode())
            zf.writestr("summary.csv", self.to_csv(documents).decode("utf-8", errors="replace"))

            # Individual document JSONs
            for doc in documents:
                doc_json = json.dumps(doc, indent=2, default=str)
                fname = doc.get("original_filename", doc.get("id", "doc"))
                safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in fname)
                zf.writestr(f"per_document/{safe_name}.json", doc_json)

                # Optionally include images
                if include_images:
                    for img_key in ["file_path", "preview_path", "preprocessed_path"]:
                        img_path = doc.get(img_key)
                        if img_path and Path(img_path).exists():
                            zf.write(img_path, f"images/{safe_name}_{img_key}.jpg")

        return output.getvalue()

    # ── Helpers ───────────────────────────────────────────────────────────

    def _auto_width(self, ws) -> None:
        """Auto-fit column widths."""
        for column_cells in ws.columns:
            length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 60)
